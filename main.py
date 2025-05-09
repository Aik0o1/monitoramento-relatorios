import pandas as pd
import os
from datetime import date
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def tratar_df(df):
    df = df.drop(df.columns[2:14], axis=1)
    df = df.drop(df.columns[14:], axis=1)
    df = df.drop(df.index[0])
    df = df[~df.map(lambda x: 'Totais' in str(x)).any(axis=1)]
    # Transformando a primeira linha no cabeçalho
    df = df.set_axis(df.iloc[0], axis=1)
    df = df[1:]
    return df

def mover_e_renomear_arquivo():
    diretorio_arq_atual = "./atual"
    arq_atual = os.listdir(diretorio_arq_atual)[0]
    caminho_arq_atual = f"{diretorio_arq_atual}/{arq_atual}"
    data_atual = date.today()
    data_formatada = data_atual.strftime("%d-%m-%Y")
    diretorio_destino = "./historico-sem-mei"
    shutil.move(caminho_arq_atual, f"{diretorio_destino}/{data_formatada}.xlsx")

def formatar_diferenca(df_atual, df_antigo, diff_indices, coluna, nome_arquivo_historico):
    """Formata as diferenças entre os DataFrames para uma visualização mais clara"""
    resultados = []
    
    for idx in diff_indices:
        if idx in df_atual.index and idx in df_antigo.index:
            valor_atual = df_atual.at[idx, coluna]
            valor_antigo = df_antigo.at[idx, coluna]
            
            # Só adiciona se os valores forem diferentes
            if valor_atual != valor_antigo and not (pd.isna(valor_atual) and pd.isna(valor_antigo)):
                tipo_evento = df_atual.at[idx, 'Tipo de Evento'] if 'Tipo de Evento' in df_atual.columns else 'N/A'
                ano = df_atual.at[idx, 'ANO'] if 'ANO' in df_atual.columns else 'N/A'
                
                resultados.append({
                    'Arquivo Histórico': nome_arquivo_historico,
                    'ANO': ano,
                    'Mês': coluna,
                    'Tipo de Evento': tipo_evento,
                    # 'Índice': idx,
                    'Valor Atual': valor_atual,
                    'Valor Antigo': valor_antigo,
                })
    
    return resultados

def comparar():
    # Cria diretório para resultados se não existir
    if not os.path.exists("./resultados_comparacao"):
        os.makedirs("./resultados_comparacao")
    
    # Obtém o arquivo atual
    arq_atual = os.listdir("./atual")[0]
    nome_arquivo_atual = os.path.splitext(arq_atual)[0]
    df_atual = pd.read_excel(f"./atual/{arq_atual}")
    df_atual = tratar_df(df_atual)
    
    # Cria diretório histórico se não existir
    if not os.path.isdir("./historico-sem-mei"):
        os.mkdir("./historico-sem-mei")
        
    dir_historico = "./historico-sem-mei"
    arquivos_historicos = os.listdir(dir_historico)
    
    # DataFrame para acumular todas as diferenças
    todas_diferencas = pd.DataFrame()
    
    for arquivo_historico in arquivos_historicos:
        print(f"\n{'='*50}")
        print(f"Comparação com o arquivo: {arquivo_historico}")
        print(f"{'='*50}")
        
        df_antigo = pd.read_excel(f"{dir_historico}/{arquivo_historico}")
        df_antigo = tratar_df(df_antigo)
        
        # Encontra as diferenças
        diferencas_arquivo = []
        
        # Identifica as colunas que existem em ambos os DataFrames
        colunas_comuns = [col for col in df_atual.columns if col in df_antigo.columns]
        
        # Para cada coluna, verifica se há diferenças
        for coluna in colunas_comuns:
            if coluna not in ['Tipo de Evento', 'ANO']:  # Ignora colunas de metadata
                # Encontra índices onde há diferenças
                indices_com_diferencas = df_atual.index[df_atual[coluna] != df_antigo[coluna]].tolist()
                
                # Adiciona índices onde um tem valor e outro é NaN
                indices_com_diferencas.extend(df_atual.index[pd.isna(df_atual[coluna]) & ~pd.isna(df_antigo[coluna])].tolist())
                indices_com_diferencas.extend(df_atual.index[~pd.isna(df_atual[coluna]) & pd.isna(df_antigo[coluna])].tolist())
                
                # Remove duplicatas
                indices_com_diferencas = list(set(indices_com_diferencas))
                
                if indices_com_diferencas:
                    diferencas_formatadas = formatar_diferenca(df_atual, df_antigo, indices_com_diferencas, coluna, arquivo_historico)
                    diferencas_arquivo.extend(diferencas_formatadas)
        
        if diferencas_arquivo:
            # Converte para DataFrame e adiciona ao acumulador
            df_diferencas_arquivo = pd.DataFrame(diferencas_arquivo)
            todas_diferencas = pd.concat([todas_diferencas, df_diferencas_arquivo], ignore_index=True)
            
            print(f"\nDiferenças encontradas com {arquivo_historico}:")
            print(df_diferencas_arquivo.to_string(index=False))
        else:
            print("Nenhuma diferença encontrada.")
    
    if not todas_diferencas.empty:
        # Gera nome do arquivo de saída com data e nome do arquivo atual
        data_atual = date.today().strftime("%d-%m-%Y")
        caminho_saida = f"./resultados_comparacao/Comparação_{nome_arquivo_atual}_{data_atual}.xlsx"
        
        # Exporta para Excel
        todas_diferencas.to_excel(caminho_saida, index=False)
        
        # Aplica formatação
        aplicar_formatacao_excel(caminho_saida)
        
        print(f"\n{'='*50}")
        print(f"Todas as diferenças foram consolidadas em: {caminho_saida}")
        print(f"{'='*50}")
    else:
        print("\nNenhuma diferença encontrada em nenhuma das comparações.")
    
    # Move o arquivo atual para o histórico após a comparação
    # mover_e_renomear_arquivo()

def aplicar_formatacao_excel(caminho_arquivo):
    """Aplica formatação ao arquivo Excel para destacar as diferenças"""
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    
    # Define estilos
    cabecalho_estilo = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    valor_atual_estilo = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    valor_antigo_estilo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Formata cabeçalhos
    for cell in ws[1]:
        cell.fill = cabecalho_estilo
        cell.font = fonte_cabecalho
    
    # Formata as células de valores
    for row in range(2, ws.max_row + 1):
        valor_atual_cell = ws.cell(row=row, column=4)  # Coluna 'Valor Atual'
        valor_antigo_cell = ws.cell(row=row, column=5)  # Coluna 'Valor Antigo'
        
        valor_atual_cell.fill = valor_atual_estilo
        valor_antigo_cell.fill = valor_antigo_estilo
    
    # Ajusta largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva as alterações
    wb.save(caminho_arquivo)

if __name__ == "__main__":
    comparar()